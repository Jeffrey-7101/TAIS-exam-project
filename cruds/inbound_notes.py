import json
import boto3
import base64
from openpyxl import Workbook

from io import BytesIO
from decimal import Decimal


dynamodb = boto3.resource('dynamodb')
inbound_notes_table = dynamodb.Table('InboundNotes')
products_table = dynamodb.Table('Products')

def convert_dynamodb_item(item):
    """Convierte elementos de DynamoDB en tipos serializables por JSON."""
    if isinstance(item, list):
        return [convert_dynamodb_item(i) for i in item]
    if isinstance(item, dict):
        return {k: convert_dynamodb_item(v) for k, v in item.items()}
    if isinstance(item, Decimal):
        return int(item) if item % 1 == 0 else float(item)  # Decimal a int o float
    return item

def convert_decimals(obj):
    """Convierte objetos Decimal a float en una estructura de datos."""
    if isinstance(obj, list):
        return [convert_decimals(i) for i in obj]
    elif isinstance(obj, dict):
        return {k: convert_decimals(v) for k, v in obj.items()}
    elif isinstance(obj, Decimal):
        return float(obj)
    else:
        return obj

def add_inbound_note(event, context):
    """Agregar una nota de entrada."""
    body = json.loads(event['body'])
    note_id = body["NoteID"]
    product_ids = body["ProductIDs"]
    quantities = body["Quantities"]

    if len(product_ids) != len(quantities):
        return {
            "statusCode": 400,
            "body": json.dumps({"message": "The length of ProductIDs and Quantities must match."})
        }

    products = []
    total_quantity = 0
    total_price = 0

    for idx, product_id in enumerate(product_ids):
        response = products_table.get_item(Key={"ProductID": product_id})
        if "Item" not in response:
            return {
                "statusCode": 404,
                "body": json.dumps({"message": f"Product with ID {product_id} not found."})
            }
        product = response["Item"]
        product["Quantity"] = Decimal(quantities[idx])
        product["UnitPrice"] = Decimal(product["UnitPrice"])
        product["TotalPrice"] = product["Quantity"] * product["UnitPrice"]

        products.append(product)
        total_quantity += product["Quantity"]
        total_price += product["TotalPrice"]

    note = {
        "NoteID": note_id,
        "Products": products,
        "TotalQuantity": Decimal(total_quantity),
        "TotalPrice": Decimal(total_price)
    }
    inbound_notes_table.put_item(Item=note)

    return {"statusCode": 201, "body": json.dumps({"message": "Inbound note added"})}

def get_all_inbound_notes(event, context):
    """Obtener todas las notas de entrada."""
    response = inbound_notes_table.scan()
    items = response['Items']  # Obtiene todas las notas
    return {
        "statusCode": 200,
        "body": json.dumps(items, default=str)  # Convierte los datos a JSON serializable
    }
    
def get_inbound_note_data(event, context):
    """Obtener los datos de una nota de entrada en formato JSON."""
    note_id = event['pathParameters']['note_id']
    response = inbound_notes_table.get_item(Key={"NoteID": note_id})
    
    # Validar si la nota existe
    if "Item" not in response:
        return {
            "statusCode": 404,
            "body": json.dumps({"message": "Note not found"})
        }

    # Retornar los datos de la nota
    return {
        "statusCode": 200,
        "body": json.dumps(response["Item"], default=str)  # Convierte los datos a JSON serializable
    }

def get_inbound_note_xlsx(event, context):
    """Generar y devolver un archivo XLSX de una nota de entrada."""
    note_id = event['pathParameters']['note_id']
    response = inbound_notes_table.get_item(Key={"NoteID": note_id})
    
    # Validar si la nota existe
    if "Item" not in response:
        return {
            "statusCode": 404,
            "body": json.dumps({"message": "Note not found"})
        }

    note = response["Item"]
    note = convert_decimals(note)

    # Crear el archivo XLSX
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inbound Note"

    # Agregar encabezados
    sheet.cell(row=1, column=1, value="Note ID")
    sheet.cell(row=1, column=2, value=note["NoteID"])
    sheet.cell(row=2, column=1, value="Total Products")
    sheet.cell(row=2, column=2, value=note["TotalQuantity"])
    sheet.cell(row=3, column=1, value="Total Price")
    sheet.cell(row=3, column=2, value=note["TotalPrice"])
    sheet.cell(row=5, column=1, value="Product Name")
    sheet.cell(row=5, column=2, value="Quantity")
    sheet.cell(row=5, column=3, value="Unit Price")
    sheet.cell(row=5, column=4, value="Total Price")

    row = 6
    for product in note["Products"]:
        sheet.cell(row=row, column=1, value=product["Name"])
        sheet.cell(row=row, column=2, value=product["Quantity"])
        sheet.cell(row=row, column=3, value=product["UnitPrice"])
        sheet.cell(row=row, column=4, value=product["Quantity"] * product["UnitPrice"])
        row += 1

    # Guardar el archivo en un buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Codificar el XLSX en Base64
    xlsx_binary = buffer.read()
    xlsx_base64 = base64.b64encode(xlsx_binary).decode('utf-8')

    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename={note_id}.xlsx",
        },
        "body": xlsx_base64,
        "isBase64Encoded": True
    }

def delete_inbound_note(event, context):
    """Eliminar una nota de entrada."""
    note_id = event['pathParameters']['note_id']
    response = inbound_notes_table.get_item(Key={"NoteID": note_id})
    if "Item" not in response:
        return {"statusCode": 404, "body": json.dumps({"message": "Inbound note not found"})}

    inbound_notes_table.delete_item(Key={"NoteID": note_id})
    return {"statusCode": 200, "body": json.dumps({"message": "Inbound note deleted successfully"})}
